import openpyxl.drawing.image
from openpyxl import load_workbook


class InserindoDados:

    def __init__(self):
        self.disciplinas = ['LÍNGUA PORTUGUESA','MATEMÁTICA','HISTÓRIA','GEOGRAFIA','QUÍMICA','FÍSICA','LÍNGUA INGLESA','BIOLOGIA','ARTE','EDUCAÇÃO FÍSICA','FILOSOFIA','SOCIOLOGIA','PROJETO DE VIDA','INTRODUÇÃO AO MUNDO DO TRABALHO','PRÁTICAS COMUNICATIVAS E CRIATIVAS','NIVELAMENTO LÍNGUA PORTUGUESA','NIVELAMENTO MATEMÁTICA','TUTORIA','HUMANIDADES E CIÊNCIAS SOCIAIS','ESTUDOS ORIENTADOS','CIÊNCIAS DA NATUREZA E SUAS TECNOLOGIAS','NÚCLEO DE INOVAÇÃO MATEMÁTICA','PESQUISA E INTERVENÇÃO','PRÁTICAS EXPERIMENTAIS','TECNOLOGIA E INOVAÇÃO','ELETIVA 1','ELETIVA 2']
        self.disciplinas_regular = ['LÍNGUA PORTUGUESA','MATEMÁTICA','HISTÓRIA','GEOGRAFIA','QUÍMICA','FÍSICA','LÍNGUA INGLESA','BIOLOGIA','ARTE','EDUCAÇÃO FÍSICA','FILOSOFIA','SOCIOLOGIA','PROJETO DE VIDA','PRÁTICAS COMUNICATIVAS E CRIATIVAS','ELETIVA 1']
        self.nome_turma = ''

        self.img1 = openpyxl.drawing.image.Image('img.png')
        self.img2 = openpyxl.drawing.image.Image('img.png')


    def fazedor(self,nomes):
        book = load_workbook(f'MODELO.xlsx')
        sheet = book.active

        if len(nomes) == 2:

            sheet['B2'].value = f'{" " * 20}{nomes[0]}'
            sheet['J2'].value = f'{" " * 20}{nomes[1]}'

            sheet['B3'].value = f'{" " * 20}{self.nome_turma}'
            sheet['J3'].value = f'{" " * 20}{self.nome_turma}'

            self.img1.anchor = 'B1'
            sheet.add_image(self.img1)

            self.img2.anchor = 'J1'
            sheet.add_image(self.img2)

            book.save(f'{self.nome_turma}/{nomes[0]} e {nomes[1]}.xlsx')

        elif len(nomes) == 1:

            sheet['B2'].value = f'{" " * 20}{nomes[0]}'
            sheet['B3'].value = f'{" " * 20}{self.nome_turma}'

            self.img1.anchor = 'B1'
            sheet.add_image(self.img1)

            book.save(f'{self.nome_turma}/{nomes[0]}.xlsx')

    def escrever(self,dados,n,nomes):


        if len(nomes) == 2:
            book = load_workbook(f'{self.nome_turma}/{nomes[0]} e {nomes[1]}.xlsx')

        elif len(nomes) == 1:
            book = load_workbook(f'{self.nome_turma}/{nomes[0]}.xlsx')

        sheet = book.active

        try:
            sheet[f'C{n}'].value = dados[0]['1º BIMESTRE']
        except:
            sheet[f'C{n}'].value = '#'
        try:
            sheet[f'D{n}'].value = dados[0]['2º BIMESTRE']
        except:
            sheet[f'D{n}'].value = '#'
        try:
            sheet[f'E{n}'].value = dados[0]['3º BIMESTRE']
        except:
            sheet[f'E{n}'].value = '#'

        try:
            sheet[f'K{n}'].value = dados[1]['1º BIMESTRE']
        except:
            sheet[f'K{n}'].value = '#'
        try:
            sheet[f'L{n}'].value = dados[1]['2º BIMESTRE']
        except:
            sheet[f'L{n}'].value = '#'
        try:
            sheet[f'M{n}'].value = dados[1]['3º BIMESTRE']
        except:
            sheet[f'M{n}'].value = '#'


        if len(nomes) == 2:
            book.save(f'{self.nome_turma}/{nomes[0]} e {nomes[1]}.xlsx')

        elif len(nomes) == 1:
            book.save(f'{self.nome_turma}/{nomes[0]}.xlsx')


    def selecionar_dados(self,listinha,nomes):
        n = 6

        if len(listinha) == 2:
            d1 = listinha[0]
            d2 = listinha[1]

        elif len(listinha) == 1:

            d1 = listinha[0]

        nomes = nomes

        for dis in self.disciplinas:

            try:
                aluno_1 = d1[dis]
            except:
                aluno_1 = {dis: None}

            try:
                aluno_2 = d2[dis]
            except:
                aluno_2 = None

            lm = [aluno_1,aluno_2]

            if dis == 'LÍNGUA PORTUGUESA':
                pass
                self.fazedor(nomes)

            self.escrever(lm,n,nomes)

            n += 1


    def iniciador(self,lista,nome_da_turma):

        inde = nome_da_turma.index('-')

        self.nome_turma = nome_da_turma[:inde].strip()

        n = 0
        listinha = []
        nomes = []

        for nome in lista:

            chaves = list(nome.keys())[0]

            if n == 0 or n == 1:

                dados = nome[chaves]

                listinha.append(dados)
                nomes.append(chaves)

                n += 1

                if n == 2:
                    self.selecionar_dados(listinha,nomes)
                    n = 0

                    listinha = []
                    nomes = []

        if len(listinha) == 1:
            self.selecionar_dados(listinha, nomes)







